from array import *
from openpyxl import *
from openpyxl.cell import *
from collections import *
import os
import glob


# Verifica se o diretorio existe(ex 5. Power) e retorna True ou False
def check_dir(address):
   if address:
      # print("\nDirectory exists: ", address)
      return True
   else:
      return False

# TODO: pegar arquivos .xlsm também
# Retorna lista de arquivos .xlsx dentro do endereço recebido
def file_exists(address):
   new_address = address + '\\*.xlsx'
   new_addres_xlsm = address + '\\*.xlsm'

   # retorna lista de arquivos .xlsx
   result_file = glob.glob(new_address)
   result_file.append(glob.glob(new_addres_xlsm))

   if result_file:
      # retorna lista
      return result_file
   else:
      print(".xlsx or .xlsm not found at: ", address)
      return False


# TODO terminar
# Dar prioridade à global checklist
# Verificar possiveis problemas por conta de ter checklists personalizadas dependendo do país
# Exemplo: http://collab.lge.com/main/display/SCAPVG/Mexico
def check_ft(address):

   ft_address = address + '\\1. FT'

   if check_dir(ft_address):
      # pega lista de arquivos .xlsx
      result_file = file_exists(ft_address)



   return "Pass"


# TODO: Inserir regra para REJUDGMENT
# TODO: Fazer uma maneira para retornar todos os fails ao invés de um só
# address = endereco onde as pastas estao; SWV = versao de SW
def check_pri(address, SWV, model, suffix):
   pri_address = address + '\\2. PRI_SIM_LOCK'

   if check_dir(pri_address):
      #pega lista de arquivos .xlsx
      result_file = file_exists(pri_address)

      if result_file:
         pri_address = result_file[0]

         # abre o arquivo de acordo com o endereço passado
         wb = load_workbook(filename=pri_address, data_only=True)

         sheet_list = wb.sheetnames

         #abre a primeira planilha com o resultado principal
         work_sheet = wb.get_sheet_by_name(name=sheet_list[0])

         if work_sheet['C2'].value == suffix:
            if work_sheet['C4'].value == SWV:
               #Para caso de Rejudgment, procura na celula adjacente F6
               if work_sheet['C6'].value == "OK" or work_sheet['F6'].value == "OK" or work_sheet['F6'].value == "ok":
                  if work_sheet['F10'].value == "OK" or work_sheet['F6'].value == "OK" or work_sheet['F6'].value == "ok":
                     if work_sheet['F11'].value == "OK" or work_sheet['F6'].value == "OK" or work_sheet['F6'].value == "ok":

                        # Campo de resultados está aparentemente OK
                        # Verificação da planilha de sim lock
                        work_sheet = wb.get_sheet_by_name(name=sheet_list[-2])
                        if work_sheet['C5'].value == "OK":
                           if work_sheet['C10'].value == SWV:
                              if work_sheet['F17'].value == "PASS":
                                 #percorre todas as celulas em E em busca de algum resultado fail ou NG
                                 col = work_sheet.columns[4]

                                 for cell_value in col:
                                    if cell_value.value == "FAIL" or cell_value.value == "Fail" or cell_value.value == "fail":
                                       return "FAIL at: ", sheet_list[-2], " Fail found at", cell_value.column, cell_value.row

                                 # percorre todas as celulas em F em busca de algum resultado fail ou NG
                                 col = work_sheet.columns[5]
                                 for cell_value in col:
                                    if cell_value.value == "FAIL" or cell_value.value == "Fail" or cell_value.value == "fail":
                                       return "FAIL at: ", sheet_list[-2], " Fail found at", cell_value.column, cell_value.row

                                 # para ter acesso ao restante das planilhas que ainda não foram verificadas
                                 new_sheet_list = sheet_list[1:len(sheet_list)-2]

                                 for sheet in new_sheet_list:
                                    work_sheet = wb.get_sheet_by_name(name=sheet)
                                    col = work_sheet.columns[5] #Coluna 'F'

                                    for cell_value in col:
                                       if cell_value.value == "Failed" or  cell_value.value == "Fail" or cell_value.value == "NG":
                                          #TODO: Inserir regra para REJUDGMENT
                                          return "FAIL at: " + sheet + ". Correct it and rerun to check for more errors"

                              else:
                                 return "FAIL at: ", sheet_list[-2], " Fail found at[F17]"
                           else:
                              return "FAIL at: ", sheet_list[-2], " check SWV[C10]"
                        else:
                           return "FAIL at: ", sheet_list[-2], " Result Field[C5]"

                        return "Pass"
                     else:
                        return "FAIL SIM Lock Tests found at " + sheet_list[0]
                  else:
                     return "FAIL Confirmation Result found at " + sheet_list[0]
               else:
                  return "FAIL - Result NG found at " + sheet_list[0]
            else:
               return "FAIL - WRONG SWV found at " + sheet_list[0]
         else:
            return "FAIL - WRONG Suffix found at " + sheet_list[0]


# TODO: IMPLEMENTAR COMPARAÇÃO COM PREVIOUS
# address = endereco onde as pastas estao; SWV = versao de SW
def check_power(address, SWV):
   pwr_address = address + '\\5. Power'

   if check_dir(pwr_address):
      # pega lista de arquivos .xlsx
      result_file = file_exists(pwr_address)
      if result_file:
         pwr_address = result_file[0]

         # abre o arquivo de acordo com o endereço passado
         wb = load_workbook(filename=pwr_address, data_only=True)
         sheet_ranges = wb.get_sheet_by_name(name='Checklist Report')

         # percorre na coluna K e mostra os valores
         colK = sheet_ranges.columns[10]  # 0 = A, 1 = B...10 = K
         for valueK in colK:
            # verifica se valor da celula é NG/Fail
            if valueK.value == "NG" or valueK.value == "ng" or valueK.value == "Fail" or valueK.value == "fail":
               return "Fail result found at Test Satus column!"

         # Converte valor da celula E8 para INT e verifica se ha algum fail nesta celula (valor diferente de 0)
         E8Int = int(sheet_ranges['E8'].value)
         if (E8Int != 0):
            return "Fail: " + str(sheet_ranges['E8'].value) + " error(s) found at result table"

         # Verifica versao de SW
         SWV_target = sheet_ranges['B5'].value
         if (SWV == SWV_target):
            return "Pass"
         else:
            return "Fail: SW Version does not match!"



# Procurar pela coluna 'Status', verificar se todas as células(issues) estão fechadas; Procurar pela coluna 'Model' e ver se todas as celulas pertencem ao mesmo modelo
# address = endereco onde as pastas estao;
def check_td(address):
   td_folder = address + '\\8. TD Defect Report'

   if check_dir(td_folder):
      file_list = file_exists(td_folder)
      if file_list:
         # file_list[0] = primeiro arquivo encontrado
         td_file = file_list[0]

         # abre o arquivo de acordo com o endereço passado
         wb = load_workbook(filename=td_file, data_only=True)
         current_sheet = wb.get_sheet_by_name(name='Sheet1')

         col_range = current_sheet['A1:AJ1']
         # percore a linha1
         for row in col_range:
            # Percorre a coluna, até achar o valor "Status"
            for cell in row:
               # print("At ", cell.column, cell.row, " cell value is: ", cell.value)
               if cell.value == "Status":
                  col_index = column_index_from_string(cell.column)

                  # Uma vez achada 'Status', todas as linhas abaixo sao percorrridas para checar se o resultado é válido
                  for cell_status in current_sheet.columns[col_index - 1]:
                     # print(cell_status.value) #Para Debug
                     if cell_status.value != "Status" and cell_status.value != "Closed" and cell_status.value != "Closed.Not a bug" and cell_status.value != "Closed.Withdrawn" and cell_status.value != "Closed.Deferred":
                        print("Error: Not all issues are properly closed! - ", cell_status.value)
                        return "Fail"


         for alt_row in current_sheet['A1:AJ1']:
            # Percorre a coluna, até achar o valor "Model"
            for cell in alt_row:
               if cell.value == "Model":
                  # print("Model column is at: ", cell.column, cell.row)
                  col_index = column_index_from_string(cell.column)

                  # valor fixo usado como referência para comparação (pega primeiro nome de modelo e compara com os subsequentes)
                  model_name = cell.offset(1, 0)

                  for cell_value in current_sheet.columns[col_index - 1]:
                     if model_name.value != cell_value.value:
                        if cell_value.value == "Model":
                           print("")
                        else:
                           print("Wrong value found!\n", "Model_name = ", model_name.value, "\n Cell_value= ",
                                 cell_value.value, "\n")
                           return "Fail"

         #print("All good!")
         return "Pass"
